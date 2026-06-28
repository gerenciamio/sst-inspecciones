from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os, io, json, base64
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

app = Flask(__name__)
CORS(app)

DATABASE_URL = os.environ.get('DATABASE_URL', '')

def get_db():
    if DATABASE_URL:
        import psycopg2
        import psycopg2.extras
        conn = psycopg2.connect(DATABASE_URL)
        return conn, 'pg'
    else:
        import sqlite3
        conn = sqlite3.connect('inspecciones.db')
        conn.row_factory = sqlite3.Row
        return conn, 'sqlite'

def init_db():
    conn, dbtype = get_db()
    if dbtype == 'pg':
        cur.execute('''CREATE TABLE IF NOT EXISTS hallazgos (
            id SERIAL PRIMARY KEY, visita_id INTEGER NOT NULL,
            lugar TEXT, situacion TEXT, recomendacion TEXT,
            foto_antes TEXT, foto_despues TEXT, estado TEXT DEFAULT 'pendiente',
            factor TEXT, prioridad TEXT, responsable TEXT,
            estado_acpm TEXT, fecha_ejecucion TEXT, fecha_seguimiento TEXT)''')
        for col in ['factor TEXT','prioridad TEXT','responsable TEXT','estado_acpm TEXT','fecha_ejecucion TEXT','fecha_seguimiento TEXT']:
            try:
                cur.execute(f"ALTER TABLE hallazgos ADD COLUMN {col}")
            except:
                pass
            CREATE TABLE IF NOT EXISTS visitas (
                id SERIAL PRIMARY KEY,
                cliente TEXT NOT NULL,
                fecha TEXT NOT NULL,
                participantes TEXT,
                asesora TEXT,
                creado TEXT NOT NULL
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS hallazgos (
                id SERIAL PRIMARY KEY,
                visita_id INTEGER NOT NULL,
                lugar TEXT,
                situacion TEXT,
                recomendacion TEXT,
                foto_antes TEXT,
                foto_despues TEXT,
                estado TEXT DEFAULT 'pendiente',
                    factor TEXT, prioridad TEXT, responsable TEXT,
                    estado_acpm TEXT, fecha_ejecucion TEXT, fecha_seguimiento TEXT
            )
        ''')
    else:
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS visitas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente TEXT NOT NULL,
                fecha TEXT NOT NULL,
                participantes TEXT,
                asesora TEXT,
                creado TEXT NOT NULL
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS hallazgos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                visita_id INTEGER NOT NULL,
                lugar TEXT,
                situacion TEXT,
                recomendacion TEXT,
                foto_antes TEXT,
                foto_despues TEXT,
                estado TEXT DEFAULT 'pendiente'
            )
        ''')
    conn.commit()
    cur.close()
    conn.close()

init_db()

@app.route('/')
def index():
    with open('index.html', encoding='utf-8') as f:
        return f.read()

@app.route('/admin')
def admin():
    with open('admin.html', encoding='utf-8') as f:
        return f.read()

@app.route('/cliente')
def cliente():
    with open('cliente.html', encoding='utf-8') as f:
        return f.read()

@app.route('/api/visita', methods=['POST'])
def crear_visita():
    data = request.get_json()
    cliente = data.get('cliente','').strip()
    fecha = data.get('fecha','').strip()
    if not cliente or not fecha:
        return jsonify({'error': 'Cliente y fecha requeridos'}), 400
    ahora = datetime.now().strftime('%Y-%m-%d %H:%M')
    conn, dbtype = get_db()
    cur = conn.cursor()
    if dbtype == 'pg':
        cur.execute(
            'INSERT INTO visitas (cliente, fecha, participantes, asesora, creado) VALUES (%s,%s,%s,%s,%s) RETURNING id',
            (cliente, fecha, data.get('participantes',''), data.get('asesora',''), ahora)
        )
        visita_id = cur.fetchone()[0]
        for h in data.get('hallazgos', []):
            cur.execute(
                'INSERT INTO hallazgos (visita_id, lugar, situacion, recomendacion, foto_antes, estado, factor, prioridad, responsable, estado_acpm, fecha_ejecucion, fecha_seguimiento) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                (visita_id, h.get('lugar',''), h.get('situacion',''), h.get('recomendacion',''), h.get('fotoBefore',''), 'pendiente', h.get('factor',''), h.get('prioridad',''), h.get('responsable',''), h.get('estado_acpm',''), h.get('fecha_ejecucion',''), h.get('fecha_seguimiento',''))
            )
    else:
        cur.execute(
            'INSERT INTO visitas (cliente, fecha, participantes, asesora, creado) VALUES (?,?,?,?,?)',
            (cliente, fecha, data.get('participantes',''), data.get('asesora',''), ahora)
        )
        visita_id = cur.lastrowid
        for h in data.get('hallazgos', []):
            cur.execute(
                'INSERT INTO hallazgos (visita_id, lugar, situacion, recomendacion, foto_antes, estado, factor, prioridad, responsable, estado_acpm, fecha_ejecucion, fecha_seguimiento) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)',
                (visita_id, h.get('lugar',''), h.get('situacion',''), h.get('recomendacion',''), h.get('fotoBefore',''), 'pendiente', h.get('factor',''), h.get('prioridad',''), h.get('responsable',''), h.get('estado_acpm',''), h.get('fecha_ejecucion',''), h.get('fecha_seguimiento',''))
            )
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True, 'visita_id': visita_id})

@app.route('/api/visita/<int:visita_id>/hallazgo', methods=['POST'])
def agregar_hallazgo(visita_id):
    data = request.get_json()
    conn, dbtype = get_db()
    cur = conn.cursor()
    ph = '%s' if dbtype == 'pg' else '?'
    cur.execute(
        f'INSERT INTO hallazgos (visita_id, lugar, situacion, recomendacion, foto_antes, estado, factor, prioridad, responsable, estado_acpm, fecha_ejecucion, fecha_seguimiento) VALUES ({ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph},{ph})',
        (visita_id, data.get('lugar',''), data.get('situacion',''), data.get('recomendacion',''), data.get('fotoBefore',''), 'pendiente', data.get('factor',''), data.get('prioridad',''), data.get('responsable',''), data.get('estado_acpm',''), data.get('fecha_ejecucion',''), data.get('fecha_seguimiento',''))
    )
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/visitas')
def listar_visitas():
    cliente_q = request.args.get('cliente', '')
    fecha_q = request.args.get('fecha', '')
    conn, dbtype = get_db()
    ph = '%s' if dbtype == 'pg' else '?'
    if dbtype == 'pg':
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    else:
        conn.row_factory = __import__('sqlite3').Row
        cur = conn.cursor()
    query = f'''SELECT v.*,
        COUNT(h.id) as total_hallazgos,
        SUM(CASE WHEN h.estado='cerrado' THEN 1 ELSE 0 END) as cerrados
        FROM visitas v LEFT JOIN hallazgos h ON h.visita_id=v.id
        WHERE 1=1'''
    params = []
    if cliente_q:
        query += f' AND LOWER(v.cliente) LIKE {ph}'
        params.append(f'%{cliente_q.lower()}%')
    if fecha_q:
        query += f' AND v.fecha={ph}'
        params.append(fecha_q)
    query += ' GROUP BY v.id ORDER BY v.creado DESC LIMIT 100'
    cur.execute(query, params)
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify(rows)

@app.route('/api/visita/<int:visita_id>')
def detalle_visita(visita_id):
    conn, dbtype = get_db()
    ph = '%s' if dbtype == 'pg' else '?'
    if dbtype == 'pg':
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    else:
        conn.row_factory = __import__('sqlite3').Row
        cur = conn.cursor()
    cur.execute(f'SELECT * FROM visitas WHERE id={ph}', (visita_id,))
    v = cur.fetchone()
    if not v:
        cur.close(); conn.close()
        return jsonify({'error': 'No encontrada'}), 404
    cur.execute(f'SELECT * FROM hallazgos WHERE visita_id={ph} ORDER BY id', (visita_id,))
    hs = [dict(h) for h in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({'visita': dict(v), 'hallazgos': hs})

@app.route('/api/hallazgo/<int:hallazgo_id>/despues', methods=['POST'])
def subir_despues(hallazgo_id):
    data = request.get_json()
    foto = data.get('foto_despues', '')
    conn, dbtype = get_db()
    ph = '%s' if dbtype == 'pg' else '?'
    cur = conn.cursor()
    cur.execute(f"UPDATE hallazgos SET foto_despues={ph}, estado='cerrado' WHERE id={ph}", (foto, hallazgo_id))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})

def b64_to_xl_image(b64_str, max_w=200, max_h=150):
    try:
        data = base64.b64decode(b64_str.split(',')[1])
        img = PILImage.open(io.BytesIO(data)).convert('RGB')
        img.thumbnail((max_w, max_h), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return XLImage(buf)
    except:
        return None

@app.route('/api/exportar/<int:visita_id>')
def exportar_excel(visita_id):
    conn, dbtype = get_db()
    ph = '%s' if dbtype == 'pg' else '?'
    if dbtype == 'pg':
        import psycopg2.extras
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    else:
        conn.row_factory = __import__('sqlite3').Row
        cur = conn.cursor()
    cur.execute(f'SELECT * FROM visitas WHERE id={ph}', (visita_id,))
    v = dict(cur.fetchone())
    cur.execute(f'SELECT * FROM hallazgos WHERE visita_id={ph} ORDER BY id', (visita_id,))
    hs = [dict(h) for h in cur.fetchall()]
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Informe de inspección'
    thin = Side(style='thin', color='000000')
    ALL_BORDERS = Border(top=thin, bottom=thin, left=thin, right=thin)

    def apply(cell, val, bold=False, sz=10, h='center', v_align='center', fill=None, border=False):
        cell.value = val
        cell.font = Font(name='Arial', bold=bold, size=sz)
        cell.alignment = Alignment(horizontal=h, vertical=v_align, wrap_text=True)
        if fill:
            cell.fill = PatternFill('solid', fgColor=fill)
        if border:
            cell.border = ALL_BORDERS

    for col, w in [('A',16),('B',10),('C',12),('D',21),('E',38),('F',38),('G',25),('H',12)]:
        ws.column_dimensions[col].width = w
    for i, h in enumerate([14,14,33,33,35], 1):
        ws.row_dimensions[i].height = h

    apply(ws['A1'], 'LOGO', bold=True)
    apply(ws['C1'], 'SEGURIDAD SALUD EN EL TRABAJO', bold=True)
    apply(ws['F1'], f'Fecha: {v["fecha"]}', bold=True, sz=8)
    apply(ws['C2'], 'INFORME Y SEGUIMIENTO A INSPECCIONES', bold=True)
    apply(ws['F2'], 'CODIGO: FT-SST-020', bold=True, sz=8)
    apply(ws['F3'], 'Pag 1 de 2', bold=True, sz=8)
    apply(ws['G3'], 'Versión 1', bold=True)
    apply(ws['A4'], f'Fecha inspección: {v["fecha"]}', bold=True, h='left')
    apply(ws['E4'], f'PARTICIPANTES: {v["participantes"]}', bold=True, h='left')

    gray = 'D3D1C7'
    for col, txt in [('A','LUGAR'),('B','SITUACIÓN ENCONTRADA'),('E','FOTO ANTES'),('F','FOTO DESPUÉS'),('G','RECOMENDACIÓN'),('H','ESTADO')]:
        apply(ws[f'{col}5'], txt, bold=True, fill=gray, border=True)

    for merge in ['A1:B3','C1:E1','C2:E3','F1:G1','F2:G2','F3:G3','A4:B4','E4:H4','B5:D5']:
        ws.merge_cells(merge)

    ROW_H = 140
    for i, h in enumerate(hs):
        r = 6 + i
        ws.row_dimensions[r].height = ROW_H
        apply(ws[f'A{r}'], h.get('lugar',''), border=True, h='left', v_align='top')
        apply(ws[f'B{r}'], h.get('situacion',''), border=True, h='left', v_align='top')
        apply(ws[f'G{r}'], h.get('recomendacion',''), border=True, h='left', v_align='top')
        estado = h.get('estado','pendiente')
        apply(ws[f'H{r}'], 'Cerrado' if estado=='cerrado' else 'Pendiente',
              border=True, fill='C0DD97' if estado=='cerrado' else 'FAC775')
        apply(ws[f'E{r}'], '', border=True)
        apply(ws[f'F{r}'], '', border=True)
        ws.merge_cells(f'B{r}:D{r}')
        if h.get('foto_antes'):
            img = b64_to_xl_image(h['foto_antes'])
            if img:
                img.anchor = f'E{r}'
                ws.add_image(img)
        if h.get('foto_despues'):
            img2 = b64_to_xl_image(h['foto_despues'])
            if img2:
                img2.anchor = f'F{r}'
                ws.add_image(img2)

    obs_row = 6 + len(hs)
    sig_row = obs_row + 1
    ws.row_dimensions[obs_row].height = 80
    ws.row_dimensions[sig_row].height = 50
    apply(ws[f'A{obs_row}'], 'OBSERVACIONES GENERALES', h='left', v_align='top')
    ws.merge_cells(f'A{obs_row}:H{obs_row}')
    apply(ws[f'A{sig_row}'], 'Firma de quien realizó la inspección' + ' '*35 + 'Firma de quien recibió la inspección', sz=11, h='center')
    ws.merge_cells(f'A{sig_row}:H{sig_row}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # ── EXCEL 2: BASE DE DATOS ACPM ──────────────────────────────────────────
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = 'BASE DE DATOS'
    ws2.merge_cells('A1:S1')
    ws2['A1'].value = 'MATRIZ DE SEGUIMIENTO A LOS PLANES DE ACCIÓN'
    ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws2['A1'].fill = PatternFill('solid', fgColor='1F4E79')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 30
    ws2.merge_cells('A2:S2')
    ws2.row_dimensions[2].height = 8
    headers = [(1,'SEDE'),(2,'MES'),(3,'FECHA'),(4,'FUENTE'),(5,'AREAS'),(6,'DESCRIPCION'),(7,'EVIDENCIA FOTOGRAFICA ANTES'),(8,'PLAN DE ACCIÓN SUGERIDO'),(9,'FACTOR DE RIESGO'),(10,'PRIORIDAD'),(11,'RESPONSABLE'),(12,'FECHA EJECUCION'),(13,'FECHA SEGUIMIENTO'),(17,'REGISTRO FOTOGRAFICO DESPUES'),(18,'ESTADO'),(19,'OBSERVACIONES')]
    for col, txt in headers:
        cell = ws2.cell(row=3, column=col, value=txt)
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E79')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = ALL_BORDERS
    ws2.row_dimensions[3].height = 35
    ws2.freeze_panes = 'A4'
    for col, w in {1:14,2:10,3:12,4:14,5:12,6:30,7:22,8:35,9:14,10:12,11:18,12:14,13:14,17:22,18:12,19:20}.items():
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    mes_map = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}
    try:
        mes = mes_map.get(int(v['fecha'].split('-')[1]), '')
    except:
        mes = ''
    alt_fill = PatternFill('solid', fgColor='EBF3FB')
    for row_idx, h in enumerate(hs, 4):
        ws2.row_dimensions[row_idx].height = 120
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col, val in [(1,v['cliente']),(2,mes),(3,v['fecha']),(4,'Inspección de Seguridad'),(5,h.get('lugar','')),(6,h.get('situacion','')),(8,h.get('recomendacion','')),(9,h.get('factor','')),(10,h.get('prioridad','')),(11,h.get('responsable','')),(12,h.get('fecha_ejecucion','')),(13,h.get('fecha_seguimiento','')),(18,h.get('estado_acpm','')),(19,'')]:
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = ALL_BORDERS
            if fill.fill_type: cell.fill = fill
        for col in [7, 17]:
            ws2.cell(row=row_idx, column=col).border = ALL_BORDERS
        if h.get('foto_antes'):
            img = b64_to_xl_image(h['foto_antes'], max_w=160, max_h=110)
            if img:
                img.anchor = openpyxl.utils.get_column_letter(7) + str(row_idx)
                ws2.add_image(img)
        if h.get('foto_despues'):
            img2 = b64_to_xl_image(h['foto_despues'], max_w=160, max_h=110)
            if img2:
                img2.anchor = openpyxl.utils.get_column_letter(17) + str(row_idx)
                ws2.add_image(img2)
    import zipfile
    buf1 = io.BytesIO(); wb1.save(buf1); buf1.seek(0)
    buf2 = io.BytesIO(); wb2.save(buf2); buf2.seek(0)
    zip_buf = io.BytesIO()
    fname_base = f"{v['cliente'].replace(' ','_')}_{v['fecha']}"
    with zipfile.ZipFile(zip_buf, 'w') as zf:
        zf.writestr(f'FT-SST-020_{fname_base}.xlsx', buf1.read())
        zf.writestr(f'Matriz_ACPM_{fname_base}.xlsx', buf2.read())
    zip_buf.seek(0)
    return send_file(zip_buf, as_attachment=True,
                     download_name=f'Informes_SST_{fname_base}.zip',
                     mimetype='application/zip')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
